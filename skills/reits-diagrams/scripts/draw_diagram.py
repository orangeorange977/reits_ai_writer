# -*- coding: utf-8 -*-
"""
通用工具：把一份"边列表"（方框A连向方框B，连线上可以标文字）渲染成层级框图 PNG。
不依赖 Graphviz 等外部程序，只用 matplotlib。适用于本申报材料里大部分"方框+箭头"
类型的示意图，比如：
  - 股权结构图（自然人/公司 一层层往下持股，直到项目公司/子项目）
  - 产品架构图（基金 -> 专项计划 -> 项目公司 -> 底层资产 这类链条）
  - 资金流向图（募集资金从基金流向项目公司/发起人等各方）
  - 资产重组步骤图（如果重组步骤本质是"A对B做了什么操作"的先后链条）

输入 JSON 格式（<edges.json>）：
{
  "edges": [
    {"from": "冯康", "to": "昊盟科技", "label": "90%"},
    {"from": "昊盟科技", "to": "奥飞数据", "label": "27.6%"},
    {"from": "奥飞数据", "to": "固安聚龙", "label": "100%"},
    {"from": "奥飞数据", "to": "昊盟盈科技", "label": "100%"},
    {"from": "固安聚龙", "to": "奥融科技", "label": "36%"},
    {"from": "昊盟盈科技", "to": "奥融科技", "label": "23.47%"},
    {"from": "奥融科技", "to": "奥飞数据数字智慧产业园1、2号楼", "label": "持有"},
    {"from": "奥融科技", "to": "非入池资产", "label": "持有"},
    {"from": "固安聚龙", "to": "昊盟盈科技", "label": "一致行动人", "relation": "peer", "style": "dashed", "color": "red"}
  ]
}
- "from"/"to" 是这条边的两端；默认按"上下级"处理——"from"在上一层、"to"在下一层。
- "label" 写在箭头旁边（持股比例、操作动作、"持有"这类关系描述都可以），可以不填。
- "relation" 可选，默认 "hierarchy"（上下级，参与层级计算）。填 "peer" 表示这是"平级"关系
  （比如一致行动人、关联方这种不是持股、不该被算作"下一层"的横向关系）——peer边不参与层级
  计算，画的时候是两个方框之间直接连一条线（默认不带箭头）。一个节点如果只有peer边、没有
  任何hierarchy边，会被当成顶层节点，所以平级关系的两个节点最好至少各自也有一条上下级边。
- "style" 可选，默认 "solid"（实线），填 "dashed" 画虚线。
- "color" 可选，默认 "black"，可以填 "red"、"blue" 等matplotlib认识的颜色名或"#RRGGBB"。
- 节点的上下层级由 hierarchy 边的方向自动推算（谁没有入边谁在最上面），不需要手动指定坐标。
- 同一个节点可以有多条入边和多条出边（比如被两方共同持有、或者一个主体同时持有好几个下级），
  这是边列表天然支持的，不需要额外处理；层级取所有上级里最深的那个+1。
- 同一层左右怎么排，按边在列表里第一次出现的顺序决定，想调整左右顺序就调整edges的顺序。

也支持直接传本skill自带的 `relations.xlsx`（不用先手动转成JSON，用户可以直接在Excel里编辑）。
约定：**一个工作表（sheet）对应一个申报材料章节**（sheet名建议用章节名，比如"REITs设立方案"）。
一个章节可能要画不止一张图，用"图片标题"这一列来分组——遇到"图片标题"非空的那一行，就
是一张新图的开始，这一行和它下面紧跟着的、"图片标题"仍为空的行，都属于同一张图，直到遇到
下一个非空的"图片标题"为止。表头认中文或英文（大小写不敏感）：
  - "上级" 或 "from"：这一行的上层方框（必填）
  - "下级" 或 "to"：这一行的下层方框（必填）
  - "标注" 或 "label"（可选）：箭头旁边的文字，不填就留空
  - "平级"（可选）：这一格填"是"/"平级"，表示这一行是横向的平级关系（一致行动人、关联方
    等），不参与层级计算，画成两个方框间的连线；留空按上下级处理。
  - "线型" 或 "style"（可选）：留空或填"实线"/"solid"是实线；填"虚线"/"dashed"是虚线。
  - "颜色" 或 "color"（可选）：留空默认黑色；可以填"红色"/"red"等中文或英文颜色名。
  - "二级标题"/"三级标题"/"四级标题"（可选）：这张图在申报材料模板里挂在哪个标题层级下面，
    只需要在这张图的第一行（"图片标题"那一行）填一次。**这里填的是给人看的章节路径，不是
    模板里一字不差的锚点文字**——真正插入模板用的 `anchor_text` 仍然要对照
    `scripts/read_chapter.py` 读出来的模板原文来确定，本脚本不负责这一步。
  - "图片标题" 或 "caption"（可选但强烈建议填）：这张图的图注文字（比如"图4 重组前状态"），
    同时也是分组依据和输出文件名。同一个sheet里如果这一列全部留空，整个sheet当成一张图处理
    （向后兼容只有"上级/下级/标注"三列的老sheet）。
空行会被跳过。想新增一张图：如果和已有图属于同一章节，在对应sheet里新起一个"图片标题"分组
即可；如果是新章节，新建一个sheet。

用法：
    python draw_diagram.py <edges.json> <输出png路径>          # 单图：JSON输入，输出单个PNG
    python draw_diagram.py relations.xlsx <输出目录>            # 多图：整个工作簿，每个sheet可能输出多张图的PNG
"""
import sys
import os
import re
import json
from collections import defaultdict, deque

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch, FancyArrowPatch, Rectangle
from matplotlib.font_manager import FontProperties


def _find_cjk_font():
    import matplotlib.font_manager as fm
    candidates = ['Microsoft YaHei', 'SimHei', 'SimSun', 'Noto Sans CJK SC', 'PingFang SC', 'WenQuanYi Zen Hei']
    available = {f.name for f in fm.fontManager.ttflist}
    for name in candidates:
        if name in available:
            return name
    return None


def build_layout(edges):
    nodes = []
    seen = set()
    children = defaultdict(list)   # from -> [(to, edge_dict)]   只放 hierarchy 边
    peer_edges = []                # 平级关系边，不参与层级计算
    for e in edges:
        a, b = e['from'], e['to']
        for n in (a, b):
            if n not in seen:
                seen.add(n)
                nodes.append(n)
        if e.get('relation') == 'peer':
            peer_edges.append(e)
        else:
            children[a].append((b, e))

    parents = defaultdict(list)
    indeg = defaultdict(int)
    for a, lst in children.items():
        for b, _ in lst:
            parents[b].append(a)
            indeg[b] += 1

    # 拓扑排序（Kahn），用于按依赖顺序算层级
    q = deque([n for n in nodes if indeg[n] == 0])
    topo = []
    indeg_work = dict(indeg)
    while q:
        n = q.popleft()
        topo.append(n)
        for b, _ in children[n]:
            indeg_work[b] -= 1
            if indeg_work[b] == 0:
                q.append(b)
    if len(topo) < len(nodes):
        # 边列表里有环，兜底：剩余节点按首次出现顺序接在后面，不阻塞画图
        for n in nodes:
            if n not in topo:
                topo.append(n)

    level = {}
    for n in topo:
        if not parents[n]:
            level[n] = 0
        else:
            level[n] = max(level.get(p, 0) for p in parents[n]) + 1

    by_level = defaultdict(list)
    for n in nodes:
        by_level[level[n]].append(n)

    return nodes, children, level, by_level, peer_edges


_COLOR_MAP = {
    '黑': 'black', '黑色': 'black',
    '红': 'red', '红色': 'red',
    '蓝': 'blue', '蓝色': 'blue',
    '绿': 'green', '绿色': 'green',
    '灰': 'gray', '灰色': 'gray',
    '橙': 'orange', '橙色': 'orange',
    '紫': 'purple', '紫色': 'purple',
}


def _resolve_color(c):
    if not c:
        return 'black'
    c = str(c).strip()
    return _COLOR_MAP.get(c, c)


def _resolve_style(s):
    if not s:
        return 'solid'
    s = str(s).strip()
    return 'dashed' if s in ('虚线', 'dashed', 'dash') else 'solid'


def draw(edges, out_path, box_w=2.6, box_h=0.8, h_gap=0.6, v_gap=1.6):
    nodes, children, level, by_level, peer_edges = build_layout(edges)
    if not nodes:
        raise ValueError('edges 为空，没有节点可画')
    max_level = max(level.values())
    max_width = max(len(ns) for ns in by_level.values())

    fig_w = max_width * (box_w + h_gap)
    fig_h = (max_level + 1) * (box_h + v_gap)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.invert_yaxis()
    ax.axis('off')

    font_name = _find_cjk_font()
    fp = FontProperties(family=font_name) if font_name else None
    if font_name is None:
        print('[警告] 系统里没找到常见中文字体，图里的中文可能显示为方块，建议装一个 Microsoft YaHei/SimHei。')

    pos = {}
    for lvl, ns in by_level.items():
        row_w = len(ns) * box_w + (len(ns) - 1) * h_gap
        start_x = (fig_w - row_w) / 2
        y = lvl * (box_h + v_gap) + box_h / 2 + 0.3
        for i, n in enumerate(ns):
            x = start_x + i * (box_w + h_gap) + box_w / 2
            pos[n] = (x, y)

    for n, (x, y) in pos.items():
        rect = FancyBboxPatch((x - box_w / 2, y - box_h / 2), box_w, box_h,
                               boxstyle='round,pad=0.02', linewidth=1.2,
                               edgecolor='black', facecolor='white')
        ax.add_patch(rect)
        ax.text(x, y, n, ha='center', va='center', fontsize=12, fontproperties=fp)

    for a, lst in children.items():
        if a not in pos:
            continue
        ax_, ay = pos[a]
        for b, e in lst:
            if b not in pos:
                continue
            bx, by = pos[b]
            start = (ax_, ay + box_h / 2)
            end = (bx, by - box_h / 2)
            color = _resolve_color(e.get('color'))
            linestyle = _resolve_style(e.get('style'))
            arrow = FancyArrowPatch(start, end, arrowstyle='-|>', mutation_scale=14,
                                     linewidth=1.2, color=color, linestyle=linestyle)
            ax.add_patch(arrow)
            label = e.get('label', '')
            mx, my = (start[0] + end[0]) / 2, (start[1] + end[1]) / 2
            ax.text(mx + 0.15, my, label, ha='left', va='center', fontsize=10, fontproperties=fp, color=color)

    # 平级关系（一致行动人、关联方等）：两个方框之间直接连一条线，不参与层级、不画箭头
    for e in peer_edges:
        a, b = e['from'], e['to']
        if a not in pos or b not in pos:
            continue
        ax_, ay = pos[a]
        bx, by = pos[b]
        color = _resolve_color(e.get('color'))
        linestyle = _resolve_style(e.get('style'))
        line = FancyArrowPatch((ax_, ay), (bx, by), arrowstyle='-', linewidth=1.2,
                                color=color, linestyle=linestyle, shrinkA=20, shrinkB=20)
        ax.add_patch(line)
        label = e.get('label', '')
        if label:
            mx, my = (ax_ + bx) / 2, (ay + by) / 2
            ax.text(mx, my - 0.15, label, ha='center', va='center', fontsize=10, fontproperties=fp, color=color)

    fig.savefig(out_path, dpi=200, bbox_inches='tight')
    plt.close(fig)


def _rect_edge_point(cx, cy, w, h, tx, ty):
    """从矩形中心(cx,cy)朝目标点(tx,ty)方向，求与矩形边框的交点（让连线正好停在框边上）。"""
    dx, dy = tx - cx, ty - cy
    if dx == 0 and dy == 0:
        return cx, cy
    sx = (w / 2) / abs(dx) if dx != 0 else float('inf')
    sy = (h / 2) / abs(dy) if dy != 0 else float('inf')
    s = min(sx, sy)
    return cx + dx * s, cy + dy * s


def draw_positioned(diagram, out_path, scale=100.0, dpi=200):
    """按用户在网页上摆好的坐标画框图（所见即所得，保留位置），渲染成 PNG。

    diagram = {
      "nodes": [{"id","text","x","y","w","h"}, ...],   # 屏幕像素坐标（y 向下）
      "edges": [{"from","to","label","style","color","arrow"}, ...],
      "canvas": {"w","h"}   # 可选
    }
    """
    nodes = diagram.get('nodes', []) or []
    edges = diagram.get('edges', []) or []
    if not nodes:
        raise ValueError('diagram 没有节点')

    canvas = diagram.get('canvas', {}) or {}
    W = canvas.get('w') or (max(n['x'] + n['w'] for n in nodes) + 40)
    H = canvas.get('h') or (max(n['y'] + n['h'] for n in nodes) + 40)

    fig, ax = plt.subplots(figsize=(W / scale, H / scale))
    ax.set_xlim(0, W)
    ax.set_ylim(0, H)
    ax.invert_yaxis()   # 屏幕坐标 y 向下
    ax.axis('off')

    font_name = _find_cjk_font()
    fp = FontProperties(family=font_name) if font_name else None

    pos = {}
    for n in nodes:
        x, y, w, h = n['x'], n['y'], n['w'], n['h']
        cx, cy = x + w / 2, y + h / 2
        pos[n['id']] = (cx, cy, w, h)
        ax.add_patch(Rectangle((x, y), w, h, linewidth=1.2,
                               edgecolor='black', facecolor='white'))
        ax.text(cx, cy, n.get('text', ''), ha='center', va='center',
                fontsize=12, fontproperties=fp, wrap=True)

    for e in edges:
        a = pos.get(e.get('from'))
        b = pos.get(e.get('to'))
        if not a or not b:
            continue
        p1 = _rect_edge_point(a[0], a[1], a[2], a[3], b[0], b[1])
        p2 = _rect_edge_point(b[0], b[1], b[2], b[3], a[0], a[1])
        color = _resolve_color(e.get('color'))
        ls = _resolve_style(e.get('style'))
        astyle = '-|>' if e.get('arrow', True) else '-'
        ax.add_patch(FancyArrowPatch(p1, p2, arrowstyle=astyle, mutation_scale=14,
                                     linewidth=1.2, color=color, linestyle=ls,
                                     shrinkA=0, shrinkB=0))
        label = e.get('label', '')
        if label:
            mx, my = (p1[0] + p2[0]) / 2, (p1[1] + p2[1]) / 2
            ax.text(mx + 4, my, label, ha='left', va='center',
                    fontsize=10, fontproperties=fp, color=color)

    fig.savefig(out_path, dpi=dpi, bbox_inches='tight')
    plt.close(fig)
    return out_path


def _safe_filename(name, fallback):
    name = (name or '').strip() or fallback
    return re.sub(r'[\\/:*?"<>|]', '_', name)


def parse_sheet_diagrams(ws):
    """把一个sheet解析成若干张图（按"图片标题"分组）。
    返回 [{'caption', 'h2', 'h3', 'h4', 'edges': [...]}, ...]；如果这个sheet没有
    "上级/下级"这两列（不是关系表），返回None。"图片标题"整列都留空时，整个sheet
    当成一张图（向后兼容只有 上级/下级/标注 三列的老sheet）。"""
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    if not rows:
        return None

    header = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

    def col_index(*names):
        for name in names:
            if name in header:
                return header.index(name)
        return None

    i_from = col_index('上级', 'from')
    i_to = col_index('下级', 'to')
    i_label = col_index('标注', 'label')
    i_relation = col_index('平级', '关系类型', 'relation')
    i_style = col_index('线型', 'style')
    i_color = col_index('颜色', 'color')
    i_h2 = col_index('二级标题', 'h2')
    i_h3 = col_index('三级标题', 'h3')
    i_h4 = col_index('四级标题', 'h4')
    i_caption = col_index('图片标题', 'caption')
    if i_from is None or i_to is None:
        return None

    def cell(row, i):
        if i is None or i >= len(row) or row[i] is None:
            return ''
        return str(row[i]).strip()

    groups = []
    current = None
    for row in rows[1:]:
        caption = cell(row, i_caption)
        if caption or current is None:
            current = {
                'caption': caption,
                'h2': cell(row, i_h2), 'h3': cell(row, i_h3), 'h4': cell(row, i_h4),
                'edges': [],
            }
            groups.append(current)

        a = row[i_from] if i_from < len(row) else None
        b = row[i_to] if i_to < len(row) else None
        if a is None or b is None:
            continue
        relation_raw = cell(row, i_relation)
        relation = 'peer' if relation_raw.lower() in ('是', '平级', 'peer', 'y', 'yes') else 'hierarchy'
        current['edges'].append({
            'from': str(a).strip(),
            'to': str(b).strip(),
            'label': cell(row, i_label),
            'relation': relation,
            'style': cell(row, i_style),
            'color': cell(row, i_color),
        })

    return [g for g in groups if g['edges']]


def draw_workbook(xlsx_path, out_dir):
    """整个工作簿一次性画完：每个sheet是一个章节，章节里按"图片标题"分组画出各自的图，
    输出到 out_dir 下的 <图片标题>.png（没填图片标题的用 <sheet名>_<序号>.png 兜底）。
    同时打印每张图对应的 二级/三级/四级标题，方便确定这张图该插进模板哪个位置。"""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    os.makedirs(out_dir, exist_ok=True)
    results = []
    for sheet_name in wb.sheetnames:
        diagrams = parse_sheet_diagrams(wb[sheet_name])
        if not diagrams:
            print(f'[跳过] 工作表 {sheet_name!r} 没有识别到"上级/下级"两列，或者是空的。')
            continue
        for idx, g in enumerate(diagrams):
            fname = _safe_filename(g['caption'], f'{sheet_name}_{idx + 1}')
            out_path = os.path.join(out_dir, f'{fname}.png')
            draw(g['edges'], out_path)
            section = ' / '.join(x for x in (sheet_name, g['h2'], g['h3'], g['h4']) if x)
            print(f"已生成: {out_path}  [章节路径: {section}]  图注: {g['caption'] or '(未填)'}")
            results.append({
                'sheet': sheet_name, 'h2': g['h2'], 'h3': g['h3'], 'h4': g['h4'],
                'caption': g['caption'], 'image_path': out_path,
            })
    if not results:
        print('[警告] 整个工作簿里没有一个sheet能识别出关系数据，检查表头是否包含"上级/下级"两列。')
    return results


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    in_path, out_path = sys.argv[1], sys.argv[2]
    if in_path.lower().endswith(('.xlsx', '.xls')):
        draw_workbook(in_path, out_path)
    else:
        with open(in_path, encoding='utf-8') as f:
            data = json.load(f)
        draw(data['edges'], out_path)
        print(f'已生成: {out_path}')


if __name__ == '__main__':
    main()
