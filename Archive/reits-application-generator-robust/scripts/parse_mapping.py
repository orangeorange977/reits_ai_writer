#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
解析REITs勾稽关系表Excel，输出结构化JSON。
用法: python parse_mapping.py <excel_path> --output <output_json>
"""

import argparse
import json
import re
import sys
import os

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_source_refs(mapping_text):
    """从勾稽列文本中提取证明材料引用编号，如 [3 ...] -> 3"""
    if not mapping_text:
        return []
    refs = re.findall(r'\[(\d+)\s', mapping_text)
    return [int(r) for r in refs]


def determine_fill_strategy(mapping_text):
    """根据勾稽描述确定填充策略"""
    if not mapping_text or mapping_text.strip() == '-':
        return 'manual'

    text = mapping_text

    has_auto = any(kw in text for kw in [
        '勾稽', 'AI检索', '提取', '审计报告', '评估报告',
        '信用记录', '承诺函', '不动产权证', '营业执照',
        '投资管理手续', '法律意见', '表格内容'
    ])
    has_template = any(kw in text for kw in [
        '模板化', '可模板', '文字内容可模板'
    ])
    has_manual = any(kw in text for kw in [
        '主观填写', '主观', '人工', '难以模板化'
    ])

    if has_auto and has_manual:
        return 'mixed'
    if has_auto and has_template:
        return 'mixed'
    if has_auto:
        return 'auto_extract'
    if has_template:
        return 'template_fill'
    if has_manual:
        return 'manual'
    return 'mixed'


def parse_table_templates(f_text):
    """解析F列固定表格模板，返回表格名称列表"""
    if not f_text or f_text.strip() == '-':
        return []
    templates = []
    parts = re.split(r'[；;\n]', f_text)
    for part in parts:
        part = part.strip()
        if not part or part == '-':
            continue
        match = re.search(r'表(\d+)', part)
        if match:
            table_num = int(match.group(1))
            name = re.sub(r'^固定表格模板[—\-—]*', '', part).strip()
            templates.append({
                'table_num': table_num,
                'name': name
            })
    return templates


def parse_mapping(excel_path):
    """解析勾稽关系表"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if '正文勾稽' in name or '勾稽' in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or all(c is None for c in row):
            continue

        col_a = str(row[0]) if row[0] else ''
        col_b = str(row[1]) if len(row) > 1 and row[1] else ''
        col_c = str(row[2]) if len(row) > 2 and row[2] else ''
        col_d = str(row[3]) if len(row) > 3 and row[3] else ''
        col_e = str(row[4]) if len(row) > 4 and row[4] else ''
        col_f = str(row[5]) if len(row) > 5 and row[5] else ''
        col_g = str(row[6]) if len(row) > 6 and row[6] else ''

        if not col_a and not col_b and not col_c and not col_d:
            continue

        section_id = ''
        level = 'subsection'
        title = ''

        if col_b and col_b.strip():
            level = 'chapter'
            section_id = col_b.strip()
            title = col_b.strip()
        elif col_c and col_c.strip():
            level = 'section'
            section_id = col_c.strip()
            title = col_c.strip()
        elif col_d and col_d.strip():
            level = 'subsection'
            section_id = col_d.strip()
            title = col_d.strip()

        if col_a and col_a.strip() in ('摘要表', '释义', '目录'):
            level = 'special'
            section_id = col_a.strip()
            title = col_a.strip()

        fill_strategy = determine_fill_strategy(col_g)
        source_refs = parse_source_refs(col_g)
        table_templates = parse_table_templates(col_f)

        entry = {
            'category': col_a.strip() if col_a.strip() else '',
            'chapter': col_b.strip() if col_b.strip() else '',
            'section': col_c.strip() if col_c.strip() else '',
            'subsection': col_d.strip() if col_d.strip() else '',
            'section_id': section_id,
            'level': level,
            'title': title,
            'content_requirement': col_e.strip() if col_e.strip() else '',
            'table_templates': table_templates,
            'mapping_rules': col_g.strip() if col_g.strip() else '',
            'source_refs': source_refs,
            'fill_strategy': fill_strategy,
        }
        entries.append(entry)

    return entries


def main():
    parser = argparse.ArgumentParser(description='解析REITs勾稽关系表')
    parser.add_argument('excel_path', help='勾稽关系表Excel路径')
    parser.add_argument('--output', '-o', required=True, help='输出JSON路径')
    parser.add_argument('--sheet', default=None, help='指定sheet名称（默认自动检测）')

    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"ERROR: Excel file not found: {args.excel_path}", file=sys.stderr)
        sys.exit(1)

    entries = parse_mapping(args.excel_path)

    output = {
        'source_file': os.path.basename(args.excel_path),
        'total_entries': len(entries),
        'entries': entries,
    }

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    strategy_counts = {}
    for e in entries:
        s = e['fill_strategy']
        strategy_counts[s] = strategy_counts.get(s, 0) + 1

    print(f"解析完成: {len(entries)} 个条目")
    print(f"填充策略分布: {strategy_counts}")
    print(f"输出文件: {args.output}")


if __name__ == '__main__':
    main()
