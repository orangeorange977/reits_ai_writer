"""
摘要表 / 释义 / 其他基本信息 数据服务

数据的唯一来源是保存文件 summary_saved.json（由用户在网页上编辑、或上传 Excel 导入后保存产生）。
没有保存过时返回空结构，等用户录入——不再自动从 docx 或 planning.md 解析。
"""
import io
import json
import logging

from openpyxl import load_workbook

from backend.services import project_store

logger = logging.getLogger(__name__)

_GROUP_KEYS = ("summary_table", "glossary", "other_info")


def saved_summary_path():
    """当前项目保存摘要表数据的文件（每个项目一份，互不覆盖）。"""
    return project_store.current_dir() / "summary_saved.json"


def save_summary_data(data: dict) -> None:
    """把网页上编辑好的摘要表/释义/其他基本信息保存到 JSON 文件。"""
    clean = {k: (data.get(k) or []) for k in _GROUP_KEYS}
    saved_summary_path().write_text(
        json.dumps(clean, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def load_saved_summary():
    """读取已保存的摘要表数据；没有则返回 None。"""
    path = saved_summary_path()
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            return {k: (data.get(k) or []) for k in _GROUP_KEYS}
        except Exception as e:
            logger.warning(f"读取已保存摘要表失败: {e}")
    return None


def get_summary_data() -> dict:
    """返回 {summary_table, glossary, other_info}。

    唯一来源是保存文件；没有保存过则返回空结构（三个空列表），由用户在网页上录入或
    Excel 导入后保存。
    """
    saved = load_saved_summary()
    if saved is not None:
        return saved
    return {"summary_table": [], "glossary": [], "other_info": []}


# Excel 三个 sheet 名 -> 结果里的键（"其他基本信息"为旧名，保留以兼容旧 Excel）
_SHEET_MAP = {
    "摘要表": "summary_table",
    "释义": "glossary",
    "项目总体情况": "other_info",
    "其他基本信息": "other_info",
}


def parse_import_excel(file_bytes: bytes) -> dict:
    """解析用户上传的 Excel：三个 sheet（摘要表/释义/其他基本信息），
    每 sheet 第一列=键、第二列=值。返回 {summary_table, glossary, other_info}。"""
    result = {"summary_table": [], "glossary": [], "other_info": []}
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    for sheet_name, result_key in _SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or all(c is None for c in row):
                continue
            label = "" if (len(row) < 1 or row[0] is None) else str(row[0]).strip()
            value = "" if (len(row) < 2 or row[1] is None) else str(row[1]).strip()
            if label == "" and value == "":
                continue
            result[result_key].append({"label": label, "value": value})

    return result
