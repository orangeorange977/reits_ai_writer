"""Excel文档解析器 - 解析电子表格数据

使用openpyxl库解析Excel(.xlsx)文件，提取工作表数据。
支持多sheet解析、表头识别和数据提取。
"""

import logging
from dataclasses import field
from typing import List, Dict, Optional
from pathlib import Path

# 复用docx_parser中的数据类
from .docx_parser import ParsedDocument, Section, TableData

logger = logging.getLogger(__name__)


# ============================================================
# 辅助函数
# ============================================================

def _get_cell_value(cell) -> str:
    """安全获取单元格值并转为字符串"""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _find_header_row(sheet, max_check_rows: int = 5) -> int:
    """识别表头行（通常是第一个有数据的行）

    Args:
        sheet: openpyxl工作表对象
        max_check_rows: 最多检查的行数

    Returns:
        表头行号（1-based），默认返回1
    """
    for row_num in range(1, min(max_check_rows + 1, sheet.max_row + 1)):
        row_values = []
        for col in range(1, sheet.max_column + 1):
            val = _get_cell_value(sheet.cell(row=row_num, column=col))
            if val:
                row_values.append(val)
        # 如果该行有多个非空值，认为是表头行
        if len(row_values) >= 2:
            return row_num
    return 1


def _sheet_to_table_data(sheet) -> TableData:
    """将工作表转换为TableData对象"""
    if sheet.max_row is None or sheet.max_row == 0:
        return TableData(headers=[], rows=[])

    header_row_num = _find_header_row(sheet)

    # 提取表头
    headers: List[str] = []
    for col in range(1, (sheet.max_column or 0) + 1):
        val = _get_cell_value(sheet.cell(row=header_row_num, column=col))
        headers.append(val)

    # 去除末尾空列
    while headers and not headers[-1]:
        headers.pop()

    num_cols = len(headers)

    # 提取数据行
    rows: List[List[str]] = []
    for row_num in range(header_row_num + 1, (sheet.max_row or 0) + 1):
        row_data: List[str] = []
        is_empty_row = True
        for col in range(1, num_cols + 1):
            val = _get_cell_value(sheet.cell(row=row_num, column=col))
            row_data.append(val)
            if val:
                is_empty_row = False
        # 跳过完全空行
        if not is_empty_row:
            rows.append(row_data)

    return TableData(headers=headers, rows=rows)


# ============================================================
# 主要解析函数
# ============================================================

def parse_xlsx(file_path: str) -> ParsedDocument:
    """主解析函数 - 解析Excel文件并返回结构化文档

    每个工作表(sheet)作为一个Section，表格数据附加在该Section中。

    Args:
        file_path: Excel文件路径

    Returns:
        ParsedDocument: 解析后的文档对象
    """
    path = Path(file_path)

    # 基本验证
    if not path.exists():
        logger.error(f"文件不存在: {file_path}")
        return ParsedDocument(
            filename=path.name,
            file_path=str(path),
            file_size=0,
            total_pages=0,
            metadata={"error": "文件不存在"}
        )

    if path.suffix.lower() not in ('.xlsx', '.xls'):
        logger.error(f"不支持的文件格式: {path.suffix}")
        return ParsedDocument(
            filename=path.name,
            file_path=str(path),
            file_size=int(path.stat().st_size),
            total_pages=0,
            metadata={"error": f"不支持的文件格式: {path.suffix}"}
        )

    try:
        from openpyxl import load_workbook

        file_size = path.stat().st_size
        wb = load_workbook(str(path), read_only=True, data_only=True)

        logger.info(f"开始解析Excel文件: {path.name} ({file_size} bytes)")

        sections: List[Section] = []
        all_text_parts: List[str] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # 每个sheet作为一个Section
            table_data = _sheet_to_table_data(sheet)

            # 构建该sheet的文本摘要
            content_parts: List[str] = []
            content_parts.append(f"工作表: {sheet_name}")
            content_parts.append(f"行数: {sheet.max_row or 0}, 列数: {sheet.max_column or 0}")

            if table_data.headers:
                content_parts.append(f"表头: {' | '.join(table_data.headers)}")

            if table_data.rows:
                content_parts.append(f"数据行数: {len(table_data.rows)}")

            content = '\n'.join(content_parts)
            all_text_parts.append(content)

            section = Section(
                title=sheet_name,
                level=1,
                content=content,
                tables=[table_data],
                source_file=path.name,
            )
            sections.append(section)

        wb.close()

        # 构造元数据
        metadata = {
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
        }

        raw_text = '\n\n'.join(all_text_parts)

        result = ParsedDocument(
            filename=path.name,
            file_path=str(path),
            file_size=file_size,
            total_pages=len(sections),  # 每个sheet视为一页
            sections=sections,
            metadata=metadata,
            raw_text=raw_text,
        )

        logger.info(f"Excel解析完成: {path.name}, 共{len(sections)}个工作表")
        return result

    except ImportError:
        logger.error("openpyxl库未安装，请运行: pip install openpyxl")
        return ParsedDocument(
            filename=path.name,
            file_path=str(path),
            file_size=int(path.stat().st_size),
            total_pages=0,
            metadata={"error": "openpyxl库未安装"}
        )
    except Exception as e:
        logger.error(f"解析Excel文件时出错: {file_path}, 错误: {e}")
        return ParsedDocument(
            filename=path.name,
            file_path=str(path),
            file_size=int(path.stat().st_size) if path.exists() else 0,
            total_pages=0,
            metadata={"error": str(e)}
        )


def extract_sheet_data(file_path: str, sheet_name: str = None) -> List[Dict]:
    """提取指定工作表的数据为字典列表

    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称，None则使用第一个sheet

    Returns:
        字典列表，表头作为key
    """
    path = Path(file_path)

    if not path.exists():
        logger.error(f"文件不存在: {file_path}")
        return []

    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)

        # 确定要读取的sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                logger.error(f"工作表 '{sheet_name}' 不存在于 {path.name}")
                wb.close()
                return []
            sheet = wb[sheet_name]
        else:
            sheet = wb.active or wb[wb.sheetnames[0]]

        # 提取表格数据
        table_data = _sheet_to_table_data(sheet)
        wb.close()

        if not table_data.headers:
            return []

        # 转换为字典列表
        result: List[Dict] = []
        for row in table_data.rows:
            row_dict = {}
            for i, header in enumerate(table_data.headers):
                if header:  # 跳过空表头列
                    value = row[i] if i < len(row) else ""
                    row_dict[header] = value
            if row_dict:  # 跳过空行
                result.append(row_dict)

        logger.info(f"提取到{len(result)}行数据: {path.name}/{sheet_name or '默认sheet'}")
        return result

    except ImportError:
        logger.error("openpyxl库未安装")
        return []
    except Exception as e:
        logger.error(f"提取工作表数据时出错: {e}")
        return []


def list_sheets(file_path: str) -> List[str]:
    """列出Excel文件中的所有工作表名

    Args:
        file_path: Excel文件路径

    Returns:
        工作表名列表
    """
    path = Path(file_path)

    if not path.exists():
        logger.error(f"文件不存在: {file_path}")
        return []

    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        logger.info(f"Excel工作表列表: {path.name} -> {sheet_names}")
        return sheet_names

    except ImportError:
        logger.error("openpyxl库未安装")
        return []
    except Exception as e:
        logger.error(f"列出工作表时出错: {e}")
        return []
