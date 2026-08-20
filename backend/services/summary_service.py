"""
摘要表 / 释义 / 其他基本信息 数据服务

数据的唯一来源是保存文件 summary_saved.json（由用户在网页上编辑、或上传 Excel 导入后保存产生）。
没有保存过时返回空结构，等用户录入——不再自动从 docx 或 planning.md 解析。
"""
import io
import json
import logging
from pathlib import Path

from openpyxl import load_workbook

from backend.config import PROJECTS_DIR, safe_project_id

logger = logging.getLogger(__name__)

# 用户在网页上核对/编辑/导入后保存的摘要表数据（唯一可信来源），按项目隔离存放
# （workspace/projects/<项目ID>/summary_saved.json）；未传项目时用默认项目目录。


def saved_summary_path(project_id: str = None) -> Path:
    return PROJECTS_DIR / safe_project_id(project_id) / "summary_saved.json"


_GROUP_KEYS = ("summary_table", "glossary", "other_info")

# 官方2024版模板卷首摘要表的22个固定行项（与 reading/summary.md 基线一字不差）。
# 未保存过任何数据时，用它作为默认空骨架展示——打开页面即见基本表格，值留空待录入。
_DEFAULT_SUMMARY_TABLE_LABELS = [
    "项目名称",
    "行业领域",
    "资产所在地",
    "资产范围",
    "建设规模合计（万元）",
    "首次发行项目/新购入项目",
    "申报基准日",
    "不动产评估净值（万元）",
    "拟发售基金总额（万元）",
    "原始权益人及相关方认购基金比例",
    "净回收资金（万元）",
    "其中，拟用于在建项目、前期工作成熟的新建项目（含改扩建）和存量资产收购的金额（万元）",
    "拟上市场所",
    "发起人（如有）",
    "原始权益人",
    "基金管理人",
    "资产支持证券管理人",
    "律师事务所及项目主办律师",
    "会计师事务所",
    "资产评估机构",
    "税务咨询机构",
    "担任财务顾问的证券公司",
]


def default_summary_data() -> dict:
    """默认骨架：摘要表22个固定行项（值为空）；释义给一行列标题占位。"""
    return {
        "summary_table": [{"label": lb, "value": ""} for lb in _DEFAULT_SUMMARY_TABLE_LABELS],
        "glossary": [{"label": "简称", "value": "释义"}],
        "other_info": [],
    }


def save_summary_data(data: dict, project_id: str = None) -> None:
    """把网页上编辑好的摘要表/释义/其他基本信息保存到该项目的 JSON 文件。"""
    clean = {k: (data.get(k) or []) for k in _GROUP_KEYS}
    path = saved_summary_path(project_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(clean, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def load_saved_summary(project_id: str = None):
    """读取该项目已保存的摘要表数据；没有则返回 None。"""
    path = saved_summary_path(project_id)
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            return {k: (data.get(k) or []) for k in _GROUP_KEYS}
        except Exception as e:
            logger.warning(f"读取已保存摘要表失败: {e}")
    return None


def get_summary_data(project_id: str = None) -> dict:
    """返回 {summary_table, glossary, other_info}。

    唯一来源是该项目的保存文件；没有保存过、或保存的内容为空（如空表
    态下点过保存）时，对应分组回填默认骨架（摘要表22个固定行项、值为
    空），保证打开页面即见基本表格；用户可在网页上直接录入或 Excel
    导入后保存。
    """
    saved = load_saved_summary(project_id) or {}
    data = {k: (saved.get(k) or []) for k in _GROUP_KEYS}
    if not data["summary_table"]:
        data["summary_table"] = [
            {"label": lb, "value": ""} for lb in _DEFAULT_SUMMARY_TABLE_LABELS
        ]
    if not data["glossary"]:
        data["glossary"] = [{"label": "简称", "value": "释义"}]
    return data


# Excel 三个 sheet 名 -> 结果里的键
_SHEET_MAP = {
    "摘要表": "summary_table",
    "释义": "glossary",
    "其他基本信息": "other_info",
}


def parse_import_excel(file_bytes: bytes) -> dict:
    """解析用户上传的 Excel：三个 sheet（摘要表/释义/其他基本信息），
    每 sheet 第一列=键、第二列=值。返回 {summary_table, glossary, other_info}。"""
    result = {"summary_table": [], "glossary": [], "other_info": []}
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    for sheet_name, result_key in _SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row or all(c is None for c in row):
                continue
            label = "" if (len(row) < 1 or row[0] is None) else str(row[0]).strip()
            value = "" if (len(row) < 2 or row[1] is None) else str(row[1]).strip()
            if label == "" and value == "":
                continue
            result[result_key].append({"label": label, "value": value})

    return result
