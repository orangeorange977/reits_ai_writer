"""Project material -> one Markdown document package.

The business-facing unit is always one source file.  Page files are internal cache
only, so a PDF remains easy to read while individual pages can still be refined by
OCR/vision and traced precisely.
"""
from __future__ import annotations

import hashlib
import json
import logging
import re
import shutil
import subprocess
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from backend.config import PROJECTS_DIR, safe_project_id
from backend.services import materials_client

logger = logging.getLogger(__name__)

PARSER_VERSION = "2026-08-18.v3"
SUPPORTED_EXTENSIONS = {
    ".pdf", ".docx", ".xlsx", ".xlsm", ".txt", ".md", ".csv", ".json",
    ".rtf", ".jpg", ".jpeg", ".png", ".tif", ".tiff",
}


def _now() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def materials_dir(project_id: str | None) -> Path:
    return PROJECTS_DIR / safe_project_id(project_id) / "materials"


def knowledge_dir(project_id: str | None) -> Path:
    return PROJECTS_DIR / safe_project_id(project_id) / "knowledge" / "documents"


def document_id(relative_path: str) -> str:
    digest = hashlib.sha1(relative_path.encode("utf-8")).hexdigest()[:14]
    stem = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", Path(relative_path).stem).strip("-")
    return f"{stem[:36] or 'document'}-{digest}"


def _package_dir(project_id: str | None, relative_path: str) -> Path:
    return knowledge_dir(project_id) / document_id(relative_path)


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(value, encoding="utf-8")


def _write_json(path: Path, value: dict) -> None:
    _write_text(path, json.dumps(value, ensure_ascii=False, indent=2))


def _load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _safe_source(project_id: str | None, relative_path: str) -> Path:
    root = materials_dir(project_id).resolve()
    source = (root / relative_path).resolve()
    if source == root or root not in source.parents or not source.is_file():
        raise FileNotFoundError(f"找不到底稿或路径越权：{relative_path}")
    return source


def _all_files(project_id: str | None) -> list[Path]:
    root = materials_dir(project_id)
    if not root.is_dir():
        return []
    return sorted(
        (p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS),
        key=lambda p: p.relative_to(root).as_posix(),
    )


def _document_number(relative_path: str) -> str:
    match = re.search(r"(?:^|/)(\d+(?:-\d+)+)", relative_path)
    return match.group(1) if match else ""


def list_documents(project_id: str | None, required_paths: Iterable[str] | None = None) -> list[dict]:
    """List source files and their package state without parsing them."""
    root = materials_dir(project_id)
    required = set(required_paths or [])
    rows = []
    for source in _all_files(project_id):
        rel = source.relative_to(root).as_posix()
        package = _package_dir(project_id, rel)
        manifest = _load_json(package / "manifest.json")
        current_sha = ""
        stale = False
        if manifest:
            current_sha = _sha256(source)
            stale = current_sha != manifest.get("source_sha256") or manifest.get("parser_version") != PARSER_VERSION
        rows.append({
            "id": document_id(rel),
            "number": _document_number(rel),
            "filename": source.name,
            "path": rel,
            "extension": source.suffix.lower(),
            "size": source.stat().st_size,
            "required": rel in required,
            "status": "stale" if stale else (manifest.get("status") or "pending"),
            "page_count": manifest.get("page_count"),
            "native_pages": manifest.get("native_pages", 0),
            "ocr_pages": manifest.get("ocr_pages", 0),
            "vision_pages": manifest.get("vision_pages", 0),
            "placeholder_pages": manifest.get("placeholder_pages", 0),
            "updated_at": manifest.get("updated_at", ""),
        })
    return rows


def _iter_docx_blocks(document):
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield Table(child, document)


def _markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    normalized = [[(c or "").replace("|", "\\|").replace("\n", "<br>") for c in r] + [""] * (width - len(r)) for r in rows]
    header = normalized[0]
    if not any(header):
        header = [f"列{i + 1}" for i in range(width)]
    body = normalized[1:]
    return "\n".join([
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(["---"] * width) + " |",
        *("| " + " | ".join(row) + " |" for row in body),
    ])


def _parse_docx(source: Path) -> str:
    from docx import Document
    from docx.table import Table

    doc = Document(str(source))
    parts = []
    for block in _iter_docx_blocks(doc):
        if isinstance(block, Table):
            rows = []
            for row in block.rows:
                rows.append(["\n".join(p.text.strip() for p in cell.paragraphs if p.text.strip()) for cell in row.cells])
            table = _markdown_table(rows)
            if table:
                parts.append(table)
        else:
            text = block.text.strip()
            if not text:
                continue
            style = (block.style.name or "").lower() if block.style else ""
            level = 0
            m = re.search(r"heading\s*(\d+)", style)
            if m:
                level = max(1, min(6, int(m.group(1))))
            parts.append(("#" * level + " " if level else "") + text)
    return "\n\n".join(parts).strip()


_WATERMARK_PARTS = ("仅限润泽科技", "再次复印或转发无效", "REITs项目使用")


def _effective_native_text(raw: str) -> str:
    """Ignore pages whose only text layer is the project watermark."""
    lines = []
    for line in (raw or "").splitlines():
        compact = re.sub(r"\s+", "", line)
        if not compact or any(part in compact for part in _WATERMARK_PARTS):
            continue
        lines.append(line.strip())
    text = materials_client._reflow_text("\n".join(lines)).strip()
    useful = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]", "", text)
    return text if len(useful) >= 40 else ""


def _page_path(package: Path, page_number: int) -> Path:
    return package / "pages" / f"page-{page_number:04d}.md"


def _page_meta_path(package: Path, page_number: int) -> Path:
    return package / "pages" / f"page-{page_number:04d}.json"


def _page_image_path(package: Path, page_number: int) -> Path:
    return package / "pages" / f"page-{page_number:04d}.png"


def ensure_page_image(project_id: str | None, relative_path: str, page_number: int, dpi: int = 170) -> Path:
    import fitz

    source = _safe_source(project_id, relative_path)
    if source.suffix.lower() != ".pdf":
        raise ValueError("只有 PDF 支持页图像")
    package = _package_dir(project_id, relative_path)
    target = _page_image_path(package, page_number)
    if target.exists():
        return target
    doc = fitz.open(str(source))
    try:
        if page_number < 1 or page_number > doc.page_count:
            raise ValueError(f"页码超出范围：{page_number}/{doc.page_count}")
        pix = doc[page_number - 1].get_pixmap(dpi=dpi, alpha=False)
        target.parent.mkdir(parents=True, exist_ok=True)
        pix.save(str(target))
    finally:
        doc.close()
    return target


def compact_image_bytes(image_path: Path, max_dimension: int = 2000) -> bytes:
    """JPEG-compress a rendered page before sending it to a vision endpoint."""
    import io
    from PIL import Image

    image = Image.open(image_path).convert("RGB")
    if max(image.size) > max_dimension:
        image.thumbnail((max_dimension, max_dimension), Image.Resampling.LANCZOS)
    output = io.BytesIO()
    image.save(output, format="JPEG", quality=88, optimize=True)
    return output.getvalue()


def _aggregate_pdf(package: Path, manifest: dict) -> str:
    chunks = []
    image_only: list[int] = []

    summaries = _load_json(package / "extractions.json").get("summaries", [])
    if summaries:
        lines = ["## Know-how 结构化精读摘要", ""]
        for summary in summaries:
            pages = "、".join(str(page) for page in summary.get("pages", []) if page)
            lines.append(f"### {summary.get('title') or summary.get('id') or '关键事实'}")
            if pages:
                lines.append(f"> 来源页：第 {pages} 页")
            for label, value in (summary.get("facts") or {}).items():
                if str(value or "").strip():
                    lines.append(f"- **{label}**：{value}")
            if summary.get("note"):
                lines.append(f"> {summary['note']}")
            lines.append("")
        chunks.append("\n".join(lines).strip())

    def flush_image_only() -> None:
        if not image_only:
            return
        ranges = []
        start = previous = image_only[0]
        for number in image_only[1:]:
            if number == previous + 1:
                previous = number
                continue
            ranges.append(str(start) if start == previous else f"{start}-{previous}")
            start = previous = number
        ranges.append(str(start) if start == previous else f"{start}-{previous}")
        chunks.append(
            f"<!-- image-only pages: {','.join(ranges)} -->\n\n"
            f"> 扫描图像页 {','.join(ranges)} 未做全文转写；当前 Know-how 需要的关键页会在提取阶段精读。"
        )
        image_only.clear()

    for page in manifest.get("pages", []):
        number = page["page"]
        text = _page_path(package, number).read_text(encoding="utf-8", errors="ignore") if _page_path(package, number).exists() else ""
        if page.get("method") in {"placeholder", "image_only"} or not text.strip():
            image_only.append(number)
            continue
        flush_image_only()
        chunks.append(f"<!-- page: {number}; method: {page.get('method', 'unknown')} -->\n\n## 第 {number} 页\n\n{text.strip()}")
    flush_image_only()
    return "\n\n---\n\n".join(chunks).strip()


def save_extraction_summary(project_id: str | None, relative_path: str, summary_id: str,
                            title: str, pages: Iterable[int], facts: dict,
                            note: str = "") -> dict:
    """Persist a compact rule-driven reading result inside the one-file Markdown.

    This is deliberately separate from structured foundation values: if table OCR is
    noisy or a later chapter does not know the field IDs, the source's Markdown still
    contains the key facts that the vision pass successfully read, with page anchors.
    """
    source = _safe_source(project_id, relative_path)
    if source.suffix.lower() != ".pdf":
        return {}
    package = _package_dir(project_id, relative_path)
    current = get_document(project_id, relative_path)
    manifest = current.get("manifest") or {}
    payload = _load_json(package / "extractions.json")
    summaries = [item for item in payload.get("summaries", []) if item.get("id") != summary_id]
    summaries.append({
        "id": summary_id, "title": title,
        "pages": sorted(set(int(page) for page in pages if int(page) > 0)),
        "facts": {str(key): str(value) for key, value in (facts or {}).items()
                  if str(value or "").strip()},
        "note": note, "updated_at": _now(),
    })
    summaries.sort(key=lambda item: str(item.get("id", "")))
    _write_json(package / "extractions.json", {"summaries": summaries})
    header = (
        f"# {source.name}\n\n> 原始文件：`{relative_path}`  \n"
        f"> 文档标识：`{manifest['id']}`  \n> 解析版本：`{PARSER_VERSION}`  \n\n"
    )
    _write_text(package / "document.md", header + _aggregate_pdf(package, manifest) + "\n")
    return summaries[-1]


def _update_counts(manifest: dict) -> None:
    methods = [p.get("method") for p in manifest.get("pages", [])]
    manifest.update({
        "native_pages": methods.count("native_text"),
        "ocr_pages": methods.count("local_ocr"),
        "vision_pages": methods.count("vision"),
        "placeholder_pages": methods.count("placeholder") + methods.count("image_only"),
    })
    manifest["status"] = "partial" if manifest["placeholder_pages"] else "ready"


def _parse_pdf(project_id: str | None, relative_path: str, source: Path, package: Path,
               full_ocr: bool) -> dict:
    import fitz

    old = _load_json(package / "manifest.json")
    old_pages = {p.get("page"): p for p in old.get("pages", [])}
    pages = []
    doc = fitz.open(str(source))
    try:
        for idx in range(doc.page_count):
            number = idx + 1
            native = _effective_native_text(doc[idx].get_text())
            method = "native_text"
            text = native
            previous = old_pages.get(number, {})
            if previous.get("method") == "vision" and _page_path(package, number).exists():
                method = "vision"
                text = _page_path(package, number).read_text(encoding="utf-8", errors="ignore")
            elif not text and full_ocr:
                text = (materials_client.ocr_page_text(source, idx) or "").strip()
                method = "local_ocr" if text else "image_only"
            elif not text:
                method = "image_only"
                text = ""
            _write_text(_page_path(package, number), text)
            meta = {
                "page": number,
                "method": method,
                "char_count": len(text),
                "refined": method == "vision",
                "updated_at": _now(),
            }
            _write_json(_page_meta_path(package, number), meta)
            pages.append(meta)
    finally:
        doc.close()
    return {"page_count": len(pages), "pages": pages}


def _parse_xlsx(source: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(source), read_only=True, data_only=True)
    parts = []
    try:
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(values):
                    rows.append(values)
            parts.append(f"## 工作表：{ws.title}\n\n{_markdown_table(rows)}")
    finally:
        wb.close()
    return "\n\n".join(parts)


def _parse_rtf(source: Path) -> str:
    """Prefer the OS converter; retain a portable best-effort fallback."""
    converter = shutil.which("textutil")
    if converter:
        result = subprocess.run(
            [converter, "-convert", "txt", "-stdout", str(source)],
            check=False, capture_output=True,
        )
        if result.returncode == 0 and result.stdout:
            return result.stdout.decode("utf-8", errors="ignore").strip()
    raw = source.read_text(encoding="utf-8", errors="ignore")
    raw = re.sub(r"\\'([0-9a-fA-F]{2})", lambda m: bytes.fromhex(m.group(1)).decode("latin1"), raw)
    raw = re.sub(r"\\u(-?\d+)\??", lambda m: chr(int(m.group(1)) % 65536), raw)
    raw = re.sub(r"\\[a-zA-Z]+-?\d*\s?", "", raw)
    return re.sub(r"[{}]", "", raw).strip()


def _parse_image(source: Path, full_ocr: bool) -> tuple[str, list[dict]]:
    text = ""
    method = "image_only"
    if full_ocr:
        try:
            import pytesseract
            from PIL import Image
            text = (pytesseract.image_to_string(
                Image.open(source).convert("RGB"), lang="chi_sim+eng") or "").strip()
            method = "local_ocr" if text else "image_only"
        except Exception as exc:
            logger.warning("图片本地 OCR 失败 %s: %s", source, exc)
    body = text if text else (
        "<!-- image-only pages: 1 -->\n\n"
        "> 该文件为图像底稿，尚未进行全文 OCR；可在底稿库中执行“本地 OCR 补全”。"
    )
    return body, [{"page": 1, "method": method, "char_count": len(text),
                   "refined": False, "updated_at": _now()}]


def build_document(project_id: str | None, relative_path: str, full_ocr: bool = False,
                   force: bool = False) -> dict:
    """Build/update one document package and return its manifest."""
    source = _safe_source(project_id, relative_path)
    if source.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"暂不支持该文件类型：{source.suffix}")
    package = _package_dir(project_id, relative_path)
    package.mkdir(parents=True, exist_ok=True)
    sha = _sha256(source)
    current = _load_json(package / "manifest.json")
    if (not force and current.get("source_sha256") == sha
            and current.get("parser_version") == PARSER_VERSION
            and (not full_ocr or current.get("placeholder_pages", 0) == 0)):
        return current

    manifest = {
        "schema_version": "1.0",
        "parser_version": PARSER_VERSION,
        "id": document_id(relative_path),
        "source_path": relative_path,
        "filename": source.name,
        "extension": source.suffix.lower(),
        "source_sha256": sha,
        "source_size": source.stat().st_size,
        "updated_at": _now(),
        "status": "ready",
        "page_count": None,
        "pages": [],
    }
    ext = source.suffix.lower()
    if ext == ".pdf":
        manifest.update(_parse_pdf(project_id, relative_path, source, package, full_ocr))
        _update_counts(manifest)
        body = _aggregate_pdf(package, manifest)
    elif ext == ".docx":
        body = _parse_docx(source)
    elif ext in {".xlsx", ".xlsm"}:
        body = _parse_xlsx(source)
    elif ext == ".rtf":
        body = _parse_rtf(source)
    elif ext in {".jpg", ".jpeg", ".png", ".tif", ".tiff"}:
        body, pages = _parse_image(source, full_ocr)
        manifest.update({"page_count": 1, "pages": pages})
        _update_counts(manifest)
    else:
        body = source.read_text(encoding="utf-8", errors="ignore")
    header = (
        f"# {source.name}\n\n"
        f"> 原始文件：`{relative_path}`  \n"
        f"> 文档标识：`{manifest['id']}`  \n"
        f"> 解析版本：`{PARSER_VERSION}`  \n\n"
    )
    _write_text(package / "document.md", header + body.strip() + "\n")
    _write_json(package / "manifest.json", manifest)
    return manifest


def build_project(project_id: str | None, relative_paths: Iterable[str] | None = None,
                  full_ocr: bool = False, force: bool = False) -> dict:
    root = materials_dir(project_id)
    paths = list(relative_paths or [p.relative_to(root).as_posix() for p in _all_files(project_id)])
    results, errors = [], []
    for rel in paths:
        try:
            results.append(build_document(project_id, rel, full_ocr=full_ocr, force=force))
        except Exception as exc:
            logger.exception("底稿 Markdown 构建失败：%s", rel)
            errors.append({"path": rel, "error": str(exc)})
            # A failed parser must still leave a discoverable Markdown artifact so
            # business users can see which source was attempted and why it failed.
            package = _package_dir(project_id, rel)
            package.mkdir(parents=True, exist_ok=True)
            _write_text(package / "document.md", (
                f"# {Path(rel).name}\n\n> 原始文件：`{rel}`  \n"
                f"> 解析状态：失败  \n> 失败原因：{exc}\n"
            ))
            _write_json(package / "manifest.json", {
                "schema_version": "1.0", "parser_version": PARSER_VERSION,
                "id": document_id(rel), "source_path": rel, "filename": Path(rel).name,
                "extension": Path(rel).suffix.lower(), "updated_at": _now(),
                "status": "error", "error": str(exc), "pages": [],
            })
    return {"processed": len(results), "errors": errors, "documents": results}


def get_document(project_id: str | None, relative_path: str) -> dict:
    source = _safe_source(project_id, relative_path)
    package = _package_dir(project_id, relative_path)
    manifest = _load_json(package / "manifest.json")
    if (not manifest or manifest.get("parser_version") != PARSER_VERSION
            or manifest.get("source_sha256") != _sha256(source)):
        # Opening an older package is also its migration path.  This removes legacy
        # per-page placeholder prose without requiring the user to delete caches.
        manifest = build_document(project_id, relative_path, force=True)
    markdown_path = package / "document.md"
    return {
        "manifest": manifest,
        "markdown": markdown_path.read_text(encoding="utf-8", errors="ignore"),
        "source_exists": source.exists(),
    }


def read_for_generation(project_id: str | None, relative_path: str, pages: str = "",
                        anchor: str = "", query: str = "", max_chars: int = 30000) -> str:
    """Read the reusable Markdown package for a chapter agent.

    This is the generation-facing boundary that prevents every chapter from parsing
    the same source again.  Page and anchor filters select from the internal page
    markers while the business-visible artifact remains one Markdown per file.
    """
    data = get_document(project_id, relative_path)
    markdown = data.get("markdown", "")
    manifest = data.get("manifest") or {}
    if manifest.get("extension") != ".pdf":
        return markdown[:max_chars]
    chunks = re.split(r"(?=<!-- page: \d+; method: )", markdown)
    header = chunks[0]
    page_chunks = chunks[1:]
    selected = page_chunks
    if pages:
        wanted = {i + 1 for i in materials_client._parse_pages(pages, int(manifest.get("page_count") or 0))}
        selected = [chunk for chunk in page_chunks if (lambda m: m and int(m.group(1)) in wanted)(re.search(r"<!-- page: (\d+);", chunk))]
        selected = [re.sub(r"\n\n---\n\n<!-- image-only pages:[\s\S]*$", "", chunk) for chunk in selected]
    elif anchor:
        needle = re.sub(r"\s+", "", anchor)
        hit = next((idx for idx, chunk in enumerate(page_chunks) if needle in re.sub(r"\s+", "", chunk)), None)
        if hit is not None:
            selected = page_chunks[hit:hit + 5]
    elif query:
        terms = [x for x in re.split(r"[\s,，、;；]+", query) if len(x) >= 2]
        ranked = sorted(
            ((sum(term in chunk for term in terms), idx, chunk) for idx, chunk in enumerate(page_chunks)),
            reverse=True,
        )
        hits = sorted([item for item in ranked if item[0] > 0][:5], key=lambda x: x[1])
        if hits:
            selected = [item[2] for item in hits]
    result = header + "".join(selected)
    return result[:max_chars]


def refine_pdf_pages(project_id: str | None, relative_path: str, page_numbers: Iterable[int],
                     instruction: str = "") -> dict:
    """Replace selected page Markdown using the vision model; local OCR is the safe fallback."""
    source = _safe_source(project_id, relative_path)
    if source.suffix.lower() != ".pdf":
        raise ValueError("只有 PDF 支持指定页精读")
    package = _package_dir(project_id, relative_path)
    manifest = _load_json(package / "manifest.json")
    if (not manifest or manifest.get("parser_version") != PARSER_VERSION
            or manifest.get("source_sha256") != _sha256(source)):
        manifest = build_document(project_id, relative_path, force=True)
    page_map = {p["page"]: p for p in manifest.get("pages", [])}
    refined = []
    for number in sorted(set(int(n) for n in page_numbers)):
        if number not in page_map:
            raise ValueError(f"页码超出范围：{number}/{manifest.get('page_count')}")
        image_path = ensure_page_image(project_id, relative_path, number)
        method = "vision"
        try:
            from backend.services import kimi_client
            text = kimi_client.vision_page_markdown(compact_image_bytes(image_path), instruction=instruction)
        except Exception as exc:
            logger.warning("视觉精读失败，回退本地 OCR：%s", exc)
            text = materials_client.ocr_page_text(source, number - 1)
            method = "local_ocr"
        text = (text or "").strip()
        if not text:
            method = "image_only"
            text = ""
        _write_text(_page_path(package, number), text)
        page_map[number].update({
            "method": method,
            "char_count": len(text),
            "refined": method == "vision",
            "instruction": instruction.strip(),
            "updated_at": _now(),
        })
        _write_json(_page_meta_path(package, number), page_map[number])
        refined.append({"page": number, "method": method, "char_count": len(text)})
    manifest["pages"] = [page_map[n] for n in sorted(page_map)]
    manifest["updated_at"] = _now()
    _update_counts(manifest)
    _write_text(package / "document.md", (
        f"# {source.name}\n\n> 原始文件：`{relative_path}`  \n"
        f"> 文档标识：`{manifest['id']}`  \n> 解析版本：`{PARSER_VERSION}`  \n\n"
        + _aggregate_pdf(package, manifest) + "\n"
    ))
    _write_json(package / "manifest.json", manifest)
    return {"manifest": manifest, "refined": refined}


def transcribe_pdf_pages_local(project_id: str | None, relative_path: str,
                               page_numbers: Iterable[int]) -> dict:
    """Write already-selected PDF pages into the one-file Markdown with local OCR.

    This is used after a rule locator has bounded a long financial report.  Structured
    values are still read from page images by the vision model; local OCR supplies the
    readable Markdown and page anchors without sending every page to the model again.
    """
    source = _safe_source(project_id, relative_path)
    if source.suffix.lower() != ".pdf":
        raise ValueError("只有 PDF 支持指定页转写")
    package = _package_dir(project_id, relative_path)
    manifest = _load_json(package / "manifest.json")
    if (not manifest or manifest.get("parser_version") != PARSER_VERSION
            or manifest.get("source_sha256") != _sha256(source)):
        manifest = build_document(project_id, relative_path, force=True)
    page_map = {p["page"]: p for p in manifest.get("pages", [])}
    transcribed = []
    for number in sorted(set(int(n) for n in page_numbers)):
        if number not in page_map:
            continue
        # Never downgrade a complete native/vision page.
        if page_map[number].get("method") in {"native_text", "vision"}:
            continue
        text = (materials_client.ocr_page_text(source, number - 1) or "").strip()
        method = "local_ocr" if text else "image_only"
        _write_text(_page_path(package, number), text)
        page_map[number].update({
            "method": method, "char_count": len(text), "refined": False,
            "updated_at": _now(),
        })
        _write_json(_page_meta_path(package, number), page_map[number])
        transcribed.append({"page": number, "method": method, "char_count": len(text)})
    manifest["pages"] = [page_map[n] for n in sorted(page_map)]
    manifest["updated_at"] = _now()
    _update_counts(manifest)
    _write_text(package / "document.md", (
        f"# {source.name}\n\n> 原始文件：`{relative_path}`  \n"
        f"> 文档标识：`{manifest['id']}`  \n> 解析版本：`{PARSER_VERSION}`  \n\n"
        + _aggregate_pdf(package, manifest) + "\n"
    ))
    _write_json(package / "manifest.json", manifest)
    return {"manifest": manifest, "transcribed": transcribed}
